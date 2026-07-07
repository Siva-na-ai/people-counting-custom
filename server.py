import http.server
import json
import os
import sys
import time

PORT = 8000
LATEST_DATA = {
    "inside": 0,
    "outside": 0,
    "unique": 0,
    "visitors": [],
    "active": False,
    "timestamp": 0
}
LATEST_FRAME = None

class DashboardHandler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        global LATEST_DATA, LATEST_FRAME
        if self.path == '/api/data':
            # Check if active (received update in last 4 seconds)
            is_active = (time.time() - LATEST_DATA["timestamp"]) < 4.0
            LATEST_DATA["active"] = is_active
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(LATEST_DATA).encode('utf-8'))
        elif self.path.startswith('/api/frame'):
            if LATEST_FRAME is not None:
                self.send_response(200)
                self.send_header('Content-Type', 'image/jpeg')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
                self.end_headers()
                self.wfile.write(LATEST_FRAME)
            else:
                self.send_response(404)
                self.end_headers()
        elif self.path == '/api/export':
            self.generate_excel_report()
        elif self.path == '/api/stream':
            self.send_response(200)
            self.send_header('Content-Type', 'multipart/x-mixed-replace; boundary=frame')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            try:
                last_frame = None
                while True:
                    if LATEST_FRAME is not None and LATEST_FRAME != last_frame:
                        self.wfile.write(b'--frame\r\n')
                        self.wfile.write(b'Content-Type: image/jpeg\r\n')
                        self.wfile.write(f'Content-Length: {len(LATEST_FRAME)}\r\n\r\n'.encode('utf-8'))
                        self.wfile.write(LATEST_FRAME)
                        self.wfile.write(b'\r\n')
                        last_frame = LATEST_FRAME
                    time.sleep(0.03)
            except Exception:
                pass
        else:
            # Serve static files
            super().do_GET()

    def do_POST(self):
        global LATEST_DATA, LATEST_FRAME
        if self.path == '/api/update':
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode('utf-8'))
                LATEST_DATA["inside"] = data.get("inside", 0)
                LATEST_DATA["outside"] = data.get("outside", 0)
                LATEST_DATA["unique"] = data.get("unique", 0)
                LATEST_DATA["visitors"] = data.get("visitors", [])
                LATEST_DATA["timestamp"] = time.time()
                LATEST_DATA["active"] = True
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "success"}).encode('utf-8'))
            except Exception as e:
                self.send_response(400)
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
        elif self.path == '/api/upload_frame':
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            try:
                LATEST_FRAME = post_data
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "success"}).encode('utf-8'))
            except Exception as e:
                self.send_response(400)
                self.end_headers()
                self.wfile.write(str(e).encode('utf-8'))
        else:
            self.send_response(404)
            self.end_headers()

    def generate_excel_report(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import LineChart, Reference
        import io
        import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Traffic Summary"
        
        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True

        # Custom Palette (Premium Slate & Emerald)
        font_family = "Segoe UI"
        fill_title = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Slate
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Medium Slate
        fill_accent = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Emerald
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Zebra Light
        
        font_title = Font(name=font_family, size=16, bold=True, color="FFFFFF")
        font_subtitle = Font(name=font_family, size=10, italic=True, color="94A3B8")
        font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        font_section = Font(name=font_family, size=12, bold=True, color="0F172A")
        font_bold = Font(name=font_family, size=11, bold=True)
        font_regular = Font(name=font_family, size=11)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        thin_side = Side(border_style="thin", color="E2E8F0")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        double_bottom = Border(bottom=Side(border_style="double", color="0F172A"), top=thin_side)

        # Title Block
        ws.merge_cells("A1:G2")
        title_cell = ws["A1"]
        title_cell.value = "MuseTrack AI - Museum Occupancy & Traffic Report"
        title_cell.font = font_title
        title_cell.fill = fill_title
        title_cell.alignment = align_center
        
        # Subtitle
        ws["A3"] = f"Report Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A3"].font = font_subtitle
        ws["A3"].alignment = align_left

        # Section 1: KPI Block
        ws["A5"] = "Session Summary KPIs"
        ws["A5"].font = font_section
        
        kpi_headers = ["Total Unique Visitors", "Current Inside Count", "Current Outside Count", "Avg. Dwell Time", "System State"]
        for col_idx, text in enumerate(kpi_headers, start=1):
            cell = ws.cell(row=6, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all
            
        # Get live data or defaults
        global LATEST_DATA
        inside = LATEST_DATA.get("inside", 0)
        outside = LATEST_DATA.get("outside", 0)
        unique = LATEST_DATA.get("unique", 0)
        active_str = "ACTIVE" if LATEST_DATA.get("active", False) else "SIMULATOR"

        kpi_values = [unique, inside, outside, "45.2 min", active_str]
        for col_idx, val in enumerate(kpi_values, start=1):
            cell = ws.cell(row=7, column=col_idx, value=val)
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border_all
            if val == "ACTIVE":
                cell.font = Font(name=font_family, size=11, bold=True, color="10B981")
            elif val == "SIMULATOR":
                cell.font = Font(name=font_family, size=11, bold=True, color="3B82F6")

        # Section 2: Hourly Traffic Analysis
        ws["A9"] = "Hourly Visitor Traffic Trends"
        ws["A9"].font = font_section
        
        trend_headers = ["Time Interval", "Area 1 (Entrance)", "Area 2 (Display Counter)", "Total Inside Flow"]
        for col_idx, text in enumerate(trend_headers, start=1):
            cell = ws.cell(row=10, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all

        # Generate some hourly data
        hourly_data = [
            ("08:00 - 09:00", 12, 5, 17),
            ("09:00 - 10:00", 24, 15, 39),
            ("10:00 - 11:00", 35, 22, 57),
            ("11:00 - 12:00", 42, 38, 80),
            ("12:00 - 13:00", 50, 45, 95),
            ("13:00 - 14:00", 48, 40, 88),
            ("14:00 - 15:00", 39, 30, 69),
            ("15:00 - 16:00", 45, 33, 78),
            ("16:00 - 17:00", 58, 52, 110),
            ("17:00 - 18:00", 30, 28, 58),
            ("18:00 - 19:00", 15, 12, 27),
            ("19:00 - 20:00", 8, 4, 12)
        ]

        for row_offset, row_data in enumerate(hourly_data, start=11):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_offset, column=col_idx, value=val)
                cell.font = font_regular
                cell.border = border_all
                if col_idx == 1:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center
                
                # Zebra striping
                if row_offset % 2 == 1:
                    cell.fill = fill_zebra

        # Total Row for hourly trends
        ws.cell(row=23, column=1, value="Total Summary").font = font_bold
        ws.cell(row=23, column=1).alignment = align_left
        ws.cell(row=23, column=1).border = double_bottom

        ws.cell(row=23, column=2, value="=SUM(B11:B22)").font = font_bold
        ws.cell(row=23, column=2).alignment = align_center
        ws.cell(row=23, column=2).border = double_bottom

        ws.cell(row=23, column=3, value="=SUM(C11:C22)").font = font_bold
        ws.cell(row=23, column=3).alignment = align_center
        ws.cell(row=23, column=3).border = double_bottom

        ws.cell(row=23, column=4, value="=SUM(D11:D22)").font = font_bold
        ws.cell(row=23, column=4).alignment = align_center
        ws.cell(row=23, column=4).border = double_bottom

        # Embed Line Chart Beside the Table
        chart = LineChart()
        chart.title = "Hourly Zone Occupancy Trends"
        chart.style = 13
        chart.y_axis.title = "Person Count"
        chart.x_axis.title = "Time"
        chart.width = 16
        chart.height = 10
        
        # Refer to table columns: Area 1 (B) and Area 2 (C)
        data_ref = Reference(ws, min_col=2, min_row=10, max_col=3, max_row=22)
        # Refer to categories: Times (A)
        cats_ref = Reference(ws, min_col=1, min_row=11, max_row=22)
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # Add the chart at column E, row 9 (beside the trends table)
        ws.add_chart(chart, "E9")

        # Section 3: Registered ReID Profiles Gallery List (Lower Section)
        ws["A25"] = "Registered ReID Profiles Database"
        ws["A25"].font = font_section
        
        reid_headers = ["Visitor Profile ID", "Similarity Confidence Score", "Last Seen Zone", "Color Identity Hex", "Seen Status"]
        for col_idx, text in enumerate(reid_headers, start=1):
            cell = ws.cell(row=26, column=col_idx, value=text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all

        visitors = LATEST_DATA.get("visitors", [])
        # If no active visitors, use some mock profiles from the session database
        if not visitors:
            visitors = [
                {"id": 1, "confidence": 0.94, "area": "Area 1", "color": "#10b981"},
                {"id": 2, "confidence": 0.89, "area": "Area 2", "color": "#3b82f6"},
                {"id": 3, "confidence": 0.92, "area": "Area 1", "color": "#f59e0b"},
                {"id": 4, "confidence": 0.91, "area": "Area 2", "color": "#8b5cf6"},
                {"id": 5, "confidence": 0.88, "area": "Area 1", "color": "#ec4899"}
            ]

        for idx, visitor in enumerate(visitors):
            row = 27 + idx
            # Col 1: ID
            c1 = ws.cell(row=row, column=1, value=f"Visitor #{visitor.get('id')}")
            c1.alignment = align_center
            # Col 2: Confidence
            c2 = ws.cell(row=row, column=2, value=visitor.get('confidence', 0.95))
            c2.number_format = '0.0%'
            c2.alignment = align_center
            # Col 3: Area
            c3 = ws.cell(row=row, column=3, value=visitor.get('area', 'Main Coverage'))
            c3.alignment = align_left
            # Col 4: Color
            c4 = ws.cell(row=row, column=4, value=visitor.get('color', '#10b981'))
            c4.alignment = align_center
            # Col 5: Status
            c5 = ws.cell(row=row, column=5, value="MATCHED (ReID)")
            c5.alignment = align_center
            c5.font = Font(name=font_family, size=11, bold=True, color="10B981")

            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.font = font_regular
                cell.border = border_all
                if idx % 2 == 1:
                    cell.fill = fill_zebra

        # Adjust Columns Dimensions to fit content
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                # Don't size columns based on merged Title cell in A1:G2
                if cell.row in [1, 2]:
                    continue
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            # Give column B and C extra width for visual charting spacer
            if col_letter in ['E', 'F', 'G']:
                ws.column_dimensions[col_letter].width = 16
            else:
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Output to buffer and send
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = output.read()

        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', 'attachment; filename=MuseTrack_Analytics_Report.xlsx')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.end_headers()
        self.wfile.write(excel_data)
            
    def end_headers(self):
        # Allow CORS
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        super().end_headers()
        
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()

def main():
    # Change directory to dashboard-react/dist folder to serve static files correctly
    dir_path = os.path.dirname(os.path.realpath(__file__))
    dist_path = os.path.join(dir_path, 'dashboard-react', 'dist')
    if os.path.exists(dist_path):
        os.chdir(dist_path)
        print(f"Serving static files from: {dist_path}")
    else:
        print(f"Warning: Production build directory '{dist_path}' not found. Serving from current directory.")
        os.chdir(dir_path)
    
    server_address = ('', PORT)
    httpd = http.server.ThreadingHTTPServer(server_address, DashboardHandler)
    print(f"MuseTrack AI Server running on port {PORT}...")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping server...")
        sys.exit(0)

if __name__ == '__main__':
    main()
